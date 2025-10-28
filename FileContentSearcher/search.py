import os
import argparse
import openpyxl
import re
import fnmatch

def _search(pattern, content, use_regex, flags):
    """Helper function to perform a single search."""
    if use_regex:
        return re.search(pattern, content, flags)
    if (flags & re.IGNORECASE):
        return pattern.lower() in content.lower()
    return pattern in content

def check_file_conditions(content, and_patterns, or_patterns, not_patterns, use_regex, ignore_case):
    """Checks if the content satisfies all AND, OR, and NOT conditions."""
    flags = re.IGNORECASE if ignore_case else 0
    
    try:
        if not_patterns:
            for pattern in not_patterns:
                if _search(pattern, content, use_regex, flags):
                    return False

        if and_patterns:
            for pattern in and_patterns:
                if not _search(pattern, content, use_regex, flags):
                    return False

        if or_patterns:
            if not any(_search(pattern, content, use_regex, flags) for pattern in or_patterns):
                return False
        
        return True
    except re.error as e:
        print(f"Regex error: {e}")
        return False

def find_match_locations(content_lines, patterns_to_find, use_regex, ignore_case):
    """Finds all lines that match any of the given patterns and their content."""
    locations = {}
    if not patterns_to_find:
        return locations
        
    flags = re.IGNORECASE if ignore_case else 0
    
    for i, line in enumerate(content_lines):
        for pattern in patterns_to_find:
            if _search(pattern, line, use_regex, flags):
                locations[i + 1] = line.strip()
                break
    return locations

def search_in_excel(filepath, **kwargs):
    """Extracts content from an Excel file and returns a dict of matching cells to their content."""
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True)
        
        # Pass 1: Check conditions on the whole file content
        full_content_list = [str(cell.value) for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if cell.value is not None]
        full_content_str = "\n".join(full_content_list)
        if not check_file_conditions(full_content_str, **kwargs):
            return {}

        # Pass 2: Find locations and their content
        locations = {}
        positive_patterns = (kwargs.get('and_patterns') or []) + (kwargs.get('or_patterns') or [])
        if not positive_patterns:
            return {}

        flags = re.IGNORECASE if kwargs.get('ignore_case') else 0

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_content = str(cell.value)
                        if any(_search(p, cell_content, kwargs.get('use_regex'), flags) for p in positive_patterns):
                            locations[f"{sheet.title}:{cell.coordinate}"] = cell_content
        return locations
    except Exception:
        return {}

def search_in_text(filepath, **kwargs):
    """Extracts content from a text file and returns a dict of matching lines to their content."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        if not check_file_conditions(content, **kwargs):
            return {}
            
        positive_patterns = (kwargs.get('and_patterns') or []) + (kwargs.get('or_patterns') or [])
        return find_match_locations(content.splitlines(), positive_patterns, kwargs.get('use_regex'), kwargs.get('ignore_case'))
    except Exception:
        return {}

def search_files(directory, include_list, exclude_list, **kwargs):
    """Walks through a directory and searches files based on boolean conditions."""
    matching_files = {}
    print(f"Searching in '{directory}'...")
    for root, _, files in os.walk(directory):
        for file in files:
            if include_list and not any(fnmatch.fnmatch(file, pattern) for pattern in include_list):
                continue
            if exclude_list and any(fnmatch.fnmatch(file, pattern) for pattern in exclude_list):
                continue

            filepath = os.path.join(root, file)
            
            if filepath.endswith('.xlsx'):
                locations = search_in_excel(filepath, **kwargs)
            else:
                locations = search_in_text(filepath, **kwargs)
            
            if locations:
                matching_files[filepath] = locations
    
    return matching_files

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Search for files based on complex AND/OR/NOT conditions."
    )
    parser.add_argument("directory", help="The directory to search in.")
    parser.add_argument("--and", action="append", dest="and_patterns", metavar="PATTERN", help="Pattern that MUST exist (can be used multiple times).")
    parser.add_argument("--or", action="append", dest="or_patterns", metavar="PATTERN", help="Pattern where at least one MUST exist (can be used multiple times).")
    parser.add_argument("--not", action="append", dest="not_patterns", metavar="PATTERN", help="Pattern that must NOT exist (can be used multiple times).")
    parser.add_argument("-r", "--regex", action="store_true", help="Treat patterns as regular expressions.")
    parser.add_argument("-i", "--ignore-case", action="store_true", help="Perform case-insensitive search.")
    parser.add_argument("--include", help="Comma-separated list of file patterns to include (e.g., '*.py,*.txt').")
    parser.add_argument("--exclude", help="Comma-separated list of file patterns to exclude (e.g., '*.log,*.tmp').")
    
    args = parser.parse_args()

    if not (args.and_patterns or args.or_patterns):
        print("Error: You must provide at least one --and or --or pattern to search for.")
        exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Error: The 'openpyxl' library is required to search Excel files.")
        print("Please install it by running: pip install openpyxl")
        exit(1)

    if not os.path.isdir(args.directory):
        print(f"Error: Directory not found at '{args.directory}'")
        exit(1)

    include_list = args.include.split(',') if args.include else []
    exclude_list = args.exclude.split(',') if args.exclude else []

    search_kwargs = {
        "and_patterns": args.and_patterns,
        "or_patterns": args.or_patterns,
        "not_patterns": args.not_patterns,
        "use_regex": args.regex,
        "ignore_case": args.ignore_case
    }

    found_files = search_files(args.directory, include_list, exclude_list, **search_kwargs)

    if found_files:
        print("\n--- Found matching files: ---")
        for filepath, locations in found_files.items():
            print(f"\n{filepath}:")
            for loc, content in locations.items():
                print(f"  {loc}: {content}")
        print("\n--------------------------")
    else:
        print("\nNo matching files found.")
